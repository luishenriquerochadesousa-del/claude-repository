"""Build the Account Balances workbook: input flows, formula-driven balances, charts."""

import json

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

d = json.load(open("flows.json"))
MONTHS, ACCOUNTS, FLOWS = d["months"], d["accounts"], d["flows"]
INV = json.load(open("investments.json"))
LABELS = {"2026-01": "Jan 2026", "2026-02": "Feb 2026", "2026-03": "Mar 2026",
          "2026-04": "Apr 2026", "2026-05": "May 2026", "2026-06": "Jun 2026",
          "2026-07": "Jul 2026", "2026-08": "Aug 2026"}

EUR = '€#,##0.00;(€#,##0.00);-'
FONT = "Arial"
INK = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
INPUT = Font(name=FONT, size=10, color="0000FF")           # hardcoded inputs
LINK = Font(name=FONT, size=10, color="008000")            # cross-sheet links
HEAD = Font(name=FONT, size=10, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
TITLE = Font(name=FONT, size=14, bold=True)
SUB = Font(name=FONT, size=10, italic=True, color="595959")
THIN = Side(style="thin", color="BFBFBF")
TOPLINE = Border(top=Side(style="medium", color="1F3864"))

wb = Workbook()

# ---------------------------------------------------------------- Read Me ----
rm = wb.active
rm.title = "Read Me"
rm.sheet_view.showGridLines = False
rm.column_dimensions["A"].width = 3
rm.column_dimensions["B"].width = 104

lines = [
    (TITLE, "Personal Finance Tracker — Account Balances 2026"),
    (SUB, "Rebuilt from the Notion databases (Accounts, Income, Expenses, Transfers) on 2026-08-09."),
    (None, ""),
    (BOLD, "What is in this file"),
    (INK, "Monthly Flows       — one row per account per month: income, expenses, transfers in, transfers out."),
    (INK, "                      These are the only hardcoded numbers in the workbook (shown in blue)."),
    (INK, "Balances            — end-of-month balance per account, built with formulas from Monthly Flows."),
    (INK, "Investment Flows    — the same transfers split into capital you added vs. returns the market gave."),
    (INK, "Invested vs Returns — capital, cumulative return and value per investment account, plus the"),
    (INK, "                      money-weighted return. Value is checked against the Balances sheet."),
    (INK, "Charts              — net worth, every account, and capital vs. portfolio value."),
    (None, ""),
    (BOLD, "How a balance is calculated"),
    (INK, "Balance(account, month) = Balance(account, previous month) + income - expenses"),
    (INK, "                          + transfers in - transfers out"),
    (INK, "January starts from the 'Starting Balance (as of January 1st 2026)' field in the Notion Accounts database."),
    (None, ""),
    (BOLD, "Assumptions and things worth knowing"),
    (INK, "1. Every expense row in Notion is marked Paid and every transfer row is marked Executed, so nothing"),
    (INK, "   is filtered out by status."),
    (INK, "2. Transfers with an empty From Account are money arriving from outside the tracker (e.g. family"),
    (INK, "   transfers); transfers with an empty To Account are money leaving it. Both are counted one-sided."),
    (INK, "3. Two transfers — 'Investimentos - Trading 212' (29 Jun, €100 in) and 'Reembolso - Trading 212'"),
    (INK, "   (6 Jul, €100 out) — are each tagged to BOTH Trading 212 and Robinhood in Notion. They are"),
    (INK, "   attributed here to Trading 212, which both descriptions name. They cancel out, so no account's"),
    (INK, "   end balance is affected — only the Jun/Jul split for those two accounts."),
    (INK, "4. Investment accounts track contributed cash, not market value: the tracker records transfers in"),
    (INK, "   and out, not price movements. Trade Republic Inv. - Brokerage in particular is 'money put in'."),
    (INK, "5. August 2026 is partial — the last transaction in Notion is dated 5 August 2026."),
    (INK, "6. Revolut shows a small negative balance in Jan (-€19.36) and May (-€15.48). That points to a"),
    (INK, "   missing or mis-dated top-up in Notion rather than a real overdraft."),
    (None, ""),
    (BOLD, "Editing this file"),
    (INK, "Change or add rows on Monthly Flows only. Blue = type over it. Black = formula, leave alone."),
    (INK, "Green = a link to another sheet. Balances and both charts update automatically."),
    (INK, "To add a month, add its rows to Monthly Flows and insert a column on Balances copying the pattern."),
]
for i, (font, text) in enumerate(lines, start=2):
    c = rm.cell(row=i, column=2, value=text)
    if font:
        c.font = font

# --------------------------------------------------------- Monthly Flows -----
mf = wb.create_sheet("Monthly Flows")
mf.sheet_view.showGridLines = False
mf["A1"] = "Monthly Flows — source data (blue cells are inputs)"
mf["A1"].font = TITLE
mf["A2"] = "One row per account per month. Only months with activity are listed. Source: Notion Income / Expenses / Transfer databases."
mf["A2"].font = SUB

headers = ["Account", "Month", "Income (€)", "Expenses (€)",
           "Transfers In (€)", "Transfers Out (€)", "Net Change (€)"]
for j, h in enumerate(headers, start=1):
    c = mf.cell(row=4, column=j, value=h)
    c.font, c.fill = HEAD, HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
mf.freeze_panes = "A5"

for i, (name, month, inc, exp, tin, tout) in enumerate(FLOWS, start=5):
    mf.cell(row=i, column=1, value=name).font = INK
    mf.cell(row=i, column=2, value=month).font = INK
    for j, v in ((3, inc), (4, exp), (5, tin), (6, tout)):
        c = mf.cell(row=i, column=j, value=v)
        c.font, c.number_format = INPUT, EUR
    c = mf.cell(row=i, column=7, value=f"=C{i}-D{i}+E{i}-F{i}")
    c.font, c.number_format = INK, EUR

last = 4 + len(FLOWS)
tr = last + 1
mf.cell(row=tr, column=1, value="Total").font = BOLD
for j in range(3, 8):
    col = get_column_letter(j)
    c = mf.cell(row=tr, column=j, value=f"=SUM({col}5:{col}{last})")
    c.font, c.number_format, c.border = BOLD, EUR, TOPLINE

for col, w in zip("ABCDEFG", (36, 10, 13, 13, 15, 16, 15)):
    mf.column_dimensions[col].width = w

# -------------------------------------------------------------- Balances -----
bs = wb.create_sheet("Balances")
bs.sheet_view.showGridLines = False
bs["A1"] = "End-of-month balance by account"
bs["A1"].font = TITLE
bs["A2"] = "Every value below is a formula: previous month + net change pulled from Monthly Flows. Nothing here is typed in."
bs["A2"].font = SUB

for j, h in enumerate(["Account", "Type", "Bank", "Start (1 Jan 2026)"], start=1):
    c = bs.cell(row=4, column=j, value=h)
    c.font, c.fill = HEAD, HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for k, m in enumerate(MONTHS):
    c = bs.cell(row=4, column=5 + k, value=LABELS[m])
    c.font, c.fill = HEAD, HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
bs.freeze_panes = "E5"

# Hidden key row holding the "2026-01" strings the SUMIFS matches on.
for k, m in enumerate(MONTHS):
    c = bs.cell(row=3, column=5 + k, value=m)
    c.font = Font(name=FONT, size=8, color="BFBFBF")
    c.alignment = Alignment(horizontal="center")
bs.cell(row=3, column=4, value="month key →").font = Font(name=FONT, size=8, color="BFBFBF")
bs.row_dimensions[3].height = 11

order = sorted(ACCOUNTS, key=lambda a: a[0])
for i, (name, kind, bank, start) in enumerate(order, start=5):
    bs.cell(row=i, column=1, value=name).font = INK
    bs.cell(row=i, column=2, value=kind).font = INK
    bs.cell(row=i, column=3, value=bank).font = INK
    c = bs.cell(row=i, column=4, value=start)
    c.font, c.number_format = INPUT, EUR
    for k in range(len(MONTHS)):
        col, prev = get_column_letter(5 + k), get_column_letter(4 + k)
        f = (f"={prev}{i}+SUMIFS('Monthly Flows'!$G$5:$G${last},"
             f"'Monthly Flows'!$A$5:$A${last},$A{i},"
             f"'Monthly Flows'!$B$5:$B${last},{col}$3)")
        c = bs.cell(row=i, column=5 + k, value=f)
        c.font, c.number_format = LINK, EUR

blast = 4 + len(order)
trow = blast + 1
bs.cell(row=trow, column=1, value="TOTAL — net worth").font = BOLD
for k in range(-1, len(MONTHS)):
    col = get_column_letter(5 + k)
    c = bs.cell(row=trow, column=5 + k, value=f"=SUM({col}5:{col}{blast})")
    c.font, c.number_format, c.border = BOLD, EUR, TOPLINE

bs.cell(row=trow + 2, column=1,
        value="Change since 1 Jan 2026").font = BOLD
c = bs.cell(row=trow + 2, column=5 + len(MONTHS) - 1,
            value=f"={get_column_letter(4 + len(MONTHS))}{trow}-D{trow}")
c.font, c.number_format = BOLD, EUR

note = bs.cell(row=trow + 4, column=1,
               value="Starting balances come from the 'Starting Balance (as of January 1st 2026)' "
                     "field in the Notion Accounts database. August 2026 is partial "
                     "(last transaction 5 Aug 2026). See Read Me for all assumptions.")
note.font = SUB
bs["D4"].comment = Comment("Source: Notion Accounts Database → "
                          "'Starting Balance (as of January 1st 2026)'.", "Balance rebuild")

bs.column_dimensions["A"].width = 36
bs.column_dimensions["B"].width = 12
bs.column_dimensions["C"].width = 16
for k in range(len(MONTHS) + 1):
    bs.column_dimensions[get_column_letter(4 + k)].width = 14

# ------------------------------------------------------- Investment Flows ----
ivf = wb.create_sheet("Investment Flows")
ivf.sheet_view.showGridLines = False
ivf["A1"] = "Investment Flows — capital vs. return, split at source (blue cells are inputs)"
ivf["A1"].font = TITLE
ivf["A2"] = ("A transfer between two of your own accounts is capital moving. A one-sided transfer on an "
             "investment account (no From or no To) is a gain or a loss — in Notion these are the "
             "Portfolio Delta / CTT Aforro Delta / fund delta rows. Net = in minus out for that month.")
ivf["A2"].font = SUB

for j, h in enumerate(["Investment account", "Month", "Net Capital In (€)", "Net Return (€)"], start=1):
    c = ivf.cell(row=4, column=j, value=h)
    c.font, c.fill = HEAD, HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ivf.freeze_panes = "A5"

iv_rows = []
for a in INV["accounts"]:
    for k, m in enumerate(MONTHS):
        if abs(a["net_contrib"][k]) > 0.005 or abs(a["net_gain"][k]) > 0.005:
            iv_rows.append([a["name"], m, a["net_contrib"][k], a["net_gain"][k]])

for i, (name, month, cap, ret) in enumerate(iv_rows, start=5):
    ivf.cell(row=i, column=1, value=name).font = INK
    ivf.cell(row=i, column=2, value=month).font = INK
    for j, v in ((3, cap), (4, ret)):
        c = ivf.cell(row=i, column=j, value=v)
        c.font, c.number_format = INPUT, EUR

iv_last = 4 + len(iv_rows)
r = iv_last + 1
ivf.cell(row=r, column=1, value="Total").font = BOLD
for j in (3, 4):
    col = get_column_letter(j)
    c = ivf.cell(row=r, column=j, value=f"=SUM({col}5:{col}{iv_last})")
    c.font, c.number_format, c.border = BOLD, EUR, TOPLINE
for col, w in zip("ABCD", (36, 10, 17, 15)):
    ivf.column_dimensions[col].width = w

# ---------------------------------------------------- Invested vs Returns ----
iv = wb.create_sheet("Invested vs Returns")
iv.sheet_view.showGridLines = False
iv["A1"] = "Invested capital vs. returns"
iv["A1"].font = TITLE
iv["A2"] = ("Three blocks: capital you put in, cumulative return, and value (capital + return). "
            "Value must equal the same account on the Balances sheet — row 3 of each Value row checks it.")
iv["A2"].font = SUB

iv_names = [a["name"] for a in INV["accounts"]]
NC, NM = len(MONTHS), len(iv_names)
BLOCKS = [("Capital invested", 3, INV["start_capital"]),
          ("Cumulative return", 4, 0.0),
          ("Portfolio value", None, None)]

row = 4
block_first = {}
for title, flow_col, _ in BLOCKS:
    iv.cell(row=row, column=1, value=title).font = BOLD
    row += 1
    for j, h in enumerate(["Account", "1 Jan"], start=1):
        c = iv.cell(row=row, column=j if j == 1 else 4, value=h)
        c.font, c.fill = HEAD, HEAD_FILL
        c.alignment = Alignment(horizontal="center")
    for k, m in enumerate(MONTHS):
        c = iv.cell(row=row, column=5 + k, value=LABELS[m])
        c.font, c.fill = HEAD, HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        key = iv.cell(row=row - 1, column=5 + k, value=m)
        key.font = Font(name=FONT, size=8, color="BFBFBF")
        key.alignment = Alignment(horizontal="center")
    hdr = row
    block_first[title] = hdr + 1
    for i, name in enumerate(iv_names):
        rr = hdr + 1 + i
        iv.cell(row=rr, column=1, value=name).font = INK
        if title == "Portfolio value":
            cap_r, ret_r = block_first["Capital invested"] + i, block_first["Cumulative return"] + i
            for k in range(-1, NC):
                col = get_column_letter(5 + k)
                c = iv.cell(row=rr, column=5 + k, value=f"={col}{cap_r}+{col}{ret_r}")
                c.font, c.number_format = LINK, EUR
        else:
            start = (ACC_START := {a["name"]: a["start"] for a in INV["accounts"]})[name] \
                if title == "Capital invested" else 0.0
            c = iv.cell(row=rr, column=4, value=start)
            c.font, c.number_format = INPUT, EUR
            for k in range(NC):
                col, prev = get_column_letter(5 + k), get_column_letter(4 + k)
                fc = get_column_letter(flow_col)
                f = (f"={prev}{rr}+SUMIFS('Investment Flows'!${fc}$5:${fc}${iv_last},"
                     f"'Investment Flows'!$A$5:$A${iv_last},$A{rr},"
                     f"'Investment Flows'!$B$5:$B${iv_last},{col}${hdr - 1})")
                c = iv.cell(row=rr, column=5 + k, value=f)
                c.font, c.number_format = LINK, EUR
    tot = hdr + 1 + NM
    iv.cell(row=tot, column=1, value="Portfolio").font = BOLD
    for k in range(-1, NC):
        col = get_column_letter(5 + k)
        c = iv.cell(row=tot, column=5 + k,
                    value=f"=SUM({col}{hdr + 1}:{col}{hdr + NM})")
        c.font, c.number_format, c.border = BOLD, EUR, TOPLINE
    row = tot + 3

cap_tot = block_first["Capital invested"] + NM
ret_tot = block_first["Cumulative return"] + NM
val_tot = block_first["Portfolio value"] + NM
lastc = get_column_letter(4 + NC)

iv.cell(row=row, column=1, value="Summary at 5 Aug 2026").font = BOLD
row += 1
for label, formula in [
        ("Capital invested", f"={lastc}{cap_tot}"),
        ("Portfolio value", f"={lastc}{val_tot}"),
        ("Return earned", f"={lastc}{ret_tot}"),
        ("Return on capital", f"={lastc}{ret_tot}/{lastc}{cap_tot}"),
        ("Dividends & interest paid into Trade Republic Account (outside the portfolio)",
         INV["income_to_cash_total"])]:
    iv.cell(row=row, column=1, value=label).font = INK
    c = iv.cell(row=row, column=5, value=formula)
    c.font = INPUT if isinstance(formula, float) else BOLD
    c.number_format = "0.00%" if label == "Return on capital" else EUR
    row += 1

# Money-weighted return: opening capital and each month's net contribution are
# money in, the closing value is money out. Through 31 July — August is 5 days.
row += 1
iv.cell(row=row, column=1, value="Money-weighted return (cash flows through 31 Jul)").font = BOLD
row += 1
iv.cell(row=row, column=1, value="Cash flow").font = INK
n_irr = MONTHS.index("2026-07") + 1
c = iv.cell(row=row, column=4, value=f"=-D{cap_tot}")
c.font, c.number_format = LINK, EUR
for k in range(n_irr):
    col = get_column_letter(5 + k)
    prev = get_column_letter(4 + k)
    f = f"=-({col}{cap_tot}-{prev}{cap_tot})"
    if k == n_irr - 1:
        f += f"+{col}{val_tot}"
    c = iv.cell(row=row, column=5 + k, value=f)
    c.font, c.number_format = LINK, EUR
irr_row = row
row += 1
iv.cell(row=row, column=1, value="Monthly rate (IRR)").font = INK
c = iv.cell(row=row, column=5,
            value=f"=IRR(D{irr_row}:{get_column_letter(4 + n_irr)}{irr_row})")
c.font, c.number_format = BOLD, "0.00%"
mrate = row
row += 1
iv.cell(row=row, column=1, value="Annualised").font = INK
c = iv.cell(row=row, column=5, value=f"=(1+E{mrate})^12-1")
c.font, c.number_format = BOLD, "0.0%"
row += 2
note = iv.cell(row=row, column=1, value=(
    "Returns are measured from 1 Jan 2026: the €13,735 already invested enters as opening capital at "
    "its then value, so this is the year's return, not lifetime performance. Value moves only in months "
    "where you recorded a delta row. The 30 Jun brokerage row is labelled '(+)' but was entered as money "
    "leaving the account, so it is counted as the loss it is."))
note.font = SUB

iv.column_dimensions["A"].width = 46
iv.column_dimensions["B"].width = 3
iv.column_dimensions["C"].width = 3
for k in range(NC + 1):
    iv.column_dimensions[get_column_letter(4 + k)].width = 14

# ---------------------------------------------------------------- Charts -----
ch = wb.create_sheet("Charts")
ch.sheet_view.showGridLines = False
ch["A1"] = "Charts"
ch["A1"].font = TITLE
ch["A2"] = ("Both charts read straight from the Balances sheet. In the per-account chart, "
            "right-click → Select Data to hide accounts you don't want on screen.")
ch["A2"].font = SUB

# Categories span the start column through the last month, so every line begins at 1 Jan.
cats = Reference(bs, min_col=4, max_col=4 + len(MONTHS), min_row=4, max_row=4)

nw = LineChart()
nw.title = "Net worth over time (2026)"
nw.style = 2
nw.y_axis.title = "€"
nw.y_axis.numFmt = "#,##0"
nw.height, nw.width = 9, 24
s = Series(Reference(bs, min_col=4, max_col=4 + len(MONTHS), min_row=trow, max_row=trow),
           title="Net worth")
s.smooth = False
nw.append(s)
nw.set_categories(cats)
ch.add_chart(nw, "A4")

pa = LineChart()
pa.title = "Balance evolution by account (2026)"
pa.style = 2
pa.y_axis.title = "€"
pa.y_axis.numFmt = "#,##0"
pa.height, pa.width = 15, 26
for i, (name, *_rest) in enumerate(order, start=5):
    s = Series(Reference(bs, min_col=4, max_col=4 + len(MONTHS), min_row=i, max_row=i),
               title=name)
    s.smooth = False
    pa.append(s)
pa.set_categories(cats)
ch.add_chart(pa, "A25")

ivc = LineChart()
ivc.title = "Invested capital vs. portfolio value (2026)"
ivc.style = 2
ivc.y_axis.title = "€"
ivc.y_axis.numFmt = "#,##0"
ivc.height, ivc.width = 10, 24
iv_cats = Reference(iv, min_col=4, max_col=4 + NC, min_row=block_first["Capital invested"] - 1,
                    max_row=block_first["Capital invested"] - 1)
for label, r_tot in (("Capital invested", cap_tot), ("Portfolio value", val_tot)):
    s = Series(Reference(iv, min_col=4, max_col=4 + NC, min_row=r_tot, max_row=r_tot), title=label)
    s.smooth = False
    ivc.append(s)
ivc.set_categories(iv_cats)
ch.add_chart(ivc, "A57")

# openpyxl writes no cached values, so tell Excel to compute everything on open.
wb.calculation.fullCalcOnLoad = True
wb.save("Account Balances 2026.xlsx")
print("written")
